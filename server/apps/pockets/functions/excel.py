from zipfile import BadZipfile

import openpyxl
from openpyxl.utils.datetime import from_excel, to_excel

from ..constants import TransactionFile
from ..models import Transaction, TransactionCategory
from ..serializers import (
    TransactionCategorySerializer,
    TransactionCreateSerializer,
)


def check_category(category_name, request, append_categories) -> None:
    """Function for check category info."""
    if not TransactionCategory.objects.filter(
        name=category_name,
    ).exists() and category_name is not None:
        category_data = {
            'name': category_name,
        }
        category_serializer = TransactionCategorySerializer(
            data=category_data,
            context={
                'request': request,
            },
        )
        category_serializer.is_valid(raise_exception=True)
        category_serializer.save()
        append_categories.append(category_data)


def get_transaction_info(sheet, row, request, append_categories) -> None:
    """Return transaction info."""
    category_name = sheet.cell(
        row=row,
        column=TransactionFile.CATEGORY_NAME_COLUMN,
    ).value
    transaction_date = from_excel(
        sheet.cell(
            row=row,
            column=TransactionFile.TRANSACTION_DATE_COLUMN,
        ).value,
    )
    amount = sheet.cell(
        row=row,
        column=TransactionFile.AMOUNT_COLUMN,
    ).value
    transaction_type = sheet.cell(
        row=row,
        column=TransactionFile.TRANSACTION_TYPE_COLUMN,
    ).value
    check_category(category_name, request, append_categories)
    transaction_data = {
        'transaction_date': transaction_date.date(),
        'amount': int(amount),
        'transaction_type': transaction_type,
        'category': (
            None
            if category_name is None
            else TransactionCategory.objects.filter(
                name=category_name,
            ).first().id
        ),
    }
    return transaction_data


def import_operation(request) -> dict:
    """Import transactions and categories from xlxs file."""
    if request.data['file'].name.split('.')[-1] != 'xlsx':
        return {
            'message': 'Файл должен иметь расширение .xlsx',
        }
    try:
        workbook = openpyxl.reader.excel.load_workbook(
            request.data['file'],
            read_only=True,
        )
    except BadZipfile:
        return {
            'message': 'Файл является битым',
        }
    workbook.active = 0
    sheet = workbook.active
    rows = sheet.max_row
    append_categories, append_transactions = [], []
    for row in range(1, rows + 1):
        transaction_data = get_transaction_info(
            sheet,
            row,
            request,
            append_categories,
        )
        transaction_serializer = TransactionCreateSerializer(
            data=transaction_data,
            context={
                'request': request,
            },
        )
        transaction_serializer.is_valid(raise_exception=True)
        append_transactions.append(transaction_data)
    Transaction.create_transactions(append_transactions, request.user)
    data = {
        'categories': append_categories,
        'transactions': append_transactions,
    }
    workbook.close()
    workbook._archive.close()
    return data


def export_operation(queryset, filename) -> None:
    """Export transactions and categpries from xlsx file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(
        title='Transactions and Categories',
        index=0,
    )
    for index, transaction in enumerate(queryset, 1):
        sheet[f'A{index}'] = (
            transaction.category.name if transaction.category else None
        )
        sheet[f'B{index}'] = to_excel(
            transaction.transaction_date,
        )
        sheet[f'C{index}'] = transaction.amount
        sheet[f'D{index}'] = transaction.transaction_type
    workbook.save(filename=filename)
